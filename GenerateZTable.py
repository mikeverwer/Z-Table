import pandas as pd
import scipy.stats
from openpyxl import Workbook
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.styles import Border, Side, Protection
from openpyxl.utils import column_index_from_string

def generate_z_table():
    """Generate Z-table from -3.99 to 3.99 in 0.01 increments"""
    z_values = [round(z * 0.01, 2) for z in range(-399, 400)]
    probabilities = scipy.stats.norm.cdf(z_values)
    return pd.DataFrame({"Z-Value": z_values, "Probability": probabilities})
        
        
def create_named_ranges(wb, ws_title, data_length, named_cells):
    # 1. Create column ranges
    column_ranges = {
        'Z_Values': f'{ws_title}!$A$2:$A${data_length+1}',
        'Probabilities': f'{ws_title}!$B$2:$B${data_length+1}'
    }
    # 2. Create single cell references
    for name, cell_ref in named_cells.items():
        if not cell_ref or len(cell_ref) < 2:
            continue  # Skip invalid
            
        col = cell_ref[0].upper()
        row = cell_ref[1:]
        column_ranges[name] = f'{ws_title}!${col}${row}'
    # 3. Add to workbook
    for name, ref_text in column_ranges.items():
        wb.defined_names.add(DefinedName(name, attr_text=ref_text))
        
        
def style_cells(ws, cells, style):
    for cell in cells:
        ws[cell].style = style


def set_all_borders(ws, cell_range):
    """Apply thin borders to all sides of each cell in range"""
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))
    
    for row in ws[cell_range]:
        for cell in row:
            cell.border = thin_border
            
            
def set_thick_outside_borders(ws, cell_range):
    """Apply thick borders only to outer edges of the range"""
    thick = Side(style='medium')
    thin = Side(style='thin')
    
    rows = list(ws[cell_range])
    for i, row in enumerate(rows):
        for j, cell in enumerate(row):
            # Determine border style based on position
            top = thick if i == 0 else thin
            bottom = thick if i == len(rows)-1 else thin
            left = thick if j == 0 else thin
            right = thick if j == len(row)-1 else thin
            
            cell.border = Border(top=top, bottom=bottom, left=left, right=right)
        
        
def format_output_cells(ws):
    for cell in ws['B'][1:]:
        cell.number_format = '0.0000'
    two_dec = ['E1', 'E4', 'H9', 'K10']
    four_dec = ['E2', 'E5', 'E7', 'H7', 'K7', 'K8']
    for cell in two_dec:
        ws[cell].number_format = '0.00'
    for cell in four_dec:
        ws[cell].number_format = '0.0000'
    
    
def set_column_width(ws, columns, width):
    for col in columns:
        ws.column_dimensions[col].width = width
        
        
def add_functions(ws):
    # Set Initial Values
    # Z1_Input
    ws['E1'] = 0
    
    # Probability_Z1
    ws['E2'] = '=IF(Z1_Input<>"", _xlfn.XLOOKUP(Z1_Input, Z_Values, Probabilities), "")'
    # Probability_Z2
    ws['E5'] = '=IF(Z2_Input<>"", _xlfn.XLOOKUP(Z2_Input, Z_Values, Probabilities), "")'
    # P(z1 < z < z2)
    ws['E7'] = '=IF(AND(Z1_Input<>"", Z2_Input<>""), ABS(Probability_Z2 - Probability_Z1), "")'
    # Sigma_x
    ws['H7'] = ('=IF(AND(sigma<>"", sampleSize_mean<>""), IF(popSize_mean="", sigma/SQRT(sampleSize_mean), '
        '(sigma/SQRT(sampleSize_mean))*SQRT((popSize_mean-sampleSize_mean)/(popSize_mean-1))), "")')
    # z_x
    ws['H9'] = '=IF(AND(x_mean<>"", mu<>"", sigma<>""), IF(sigma_x<>"", (x_mean-mu)/sigma_x, (x_mean - mu)/sigma), "")'
    # pBar_calc
    ws['K7'] = '=IF(AND(x_prop<>"", sampleSize_prop<>""), x_prop/sampleSize_prop, "")'
    # Sigma_p
    ws['K8'] = ('=IF(AND(p<>"", sampleSize_prop<>""), IF(popSize_prop="", SQRT(p*(1-p)/sampleSize_prop), '
                'SQRT(p*(1-p)/sampleSize_prop)*SQRT((popSize_prop-sampleSize_prop)/(popSize_prop-1))), "")')
    # z_p
    ws['K10'] = '=IF(AND(sigma_p<>"", OR(x_prop<>"", pbar_man<>"")), IF(pbar_man="", (pbar_calc - p)/sigma_p, (pbar_man - p)/sigma_p), "")'
    
    
def set_readonly_cells(ws, cell_ranges, is_readonly=True):
    """
    Only locks specified cells, leaves others editable
    - cell_ranges: List of cells/ranges to protect ('A1', 'B2:C10')
    - is_readonly: True to lock, False to unlock
    """
    # First ensure ALL cells are unlocked by default
    for row in ws.iter_rows():
        for cell in row:
            cell.protection = Protection(locked=False)
    
    # Then lock only the specified cells/ranges
    for cell_range in cell_ranges:
        if ':' in cell_range:
            # Handle ranges (e.g., 'A1:B10')
            start, end = cell_range.split(':')
            min_col = column_index_from_string(start[0])
            min_row = int(start[1:])
            max_col = column_index_from_string(end[0])
            max_row = int(end[1:])
            
            for row in ws.iter_rows(min_row=min_row, max_row=max_row,
                                 min_col=min_col, max_col=max_col):
                for cell in row:
                    cell.protection = Protection(locked=is_readonly)
        else:
            # Handle single cells (e.g., 'A1')
            ws[cell_range].protection = Protection(locked=is_readonly)
    
    # Enable sheet protection (required for locking to take effect)
    ws.protection.sheet = True
    ws.protection.password = 'idgaf'
    ws.protection.enable()

def create_excel_file():
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Z_Table"
    
    # Add Z-table data
    df = generate_z_table()
    
    # Write headers
    ws['A1'] = "Z-Value"
    ws['B1'] = "Probability"
    ws['G1'] = "Sample Mean"
    ws['J1'] = "Sample Proportion"
    heading_cells: list = ['A1', 'B1', 'G1', 'J1']
    
    # Write data
    for idx, row in df.iterrows():
        ws.cell(row=idx+2, column=1, value=row['Z-Value'])
        ws.cell(row=idx+2, column=2, value=row['Probability'])
        
    named_cells:dict = {
        'Z1_Input': 'E1',
        'Probability_Z1': 'E2',
        'Z2_Input': 'E4',
        'Probability_Z2': 'E5',
        'mu': 'H2',
        'sigma': 'H3',
        'sampleSize_mean': 'H4',
        'popSize_mean': 'H5',
        'x_mean': 'H6',
        'sigma_x': 'H7',
        'p': 'K2',
        'sampleSize_prop': 'K3',
        'popSize_prop': 'K4',
        'x_prop': 'K5',
        'pbar_man': 'K6',
        'pbar_calc': 'K7',
        'sigma_p': 'K8'
    }
    create_named_ranges(wb, ws.title, len(df), named_cells)
    
    # Add cell labels
    # Z-lookup
    ws['D1'] = "Input z₁"
    ws['D2'] = "P(z < z₁)"
    ws['D4'] = "Input z₂"
    ws['D5'] = 'P(z < z₂)'
    ws['D7'] = "P(z₁ < z < z₂)"
    # sample mean
    ws['G2'] = "μ"
    ws['G3'] = "σ"
    ws['G4'] = "n"
    ws['G5'] = "N"
    ws['G6'] = "x"
    ws['G7'] = "σₓ"
    ws['G9'] = "zₓ"
    #sample proportion
    ws['J2'] = "p"
    ws['J3'] = "n"
    ws['J4'] = "N"
    ws['J5'] = "x"
    ws['J6'] = "p̄ (manual)"
    ws['J7'] = "p̄ (calc)"
    ws['J8'] = "σₚ"
    ws['J10'] = "zₚ"
    
    input_cells: list[str] = ['E1', 'E4', 'H2', 'H3', 'H4', 'H5', 'H6', 'K2', 'K3', 'K4', 'K5', 'K6']
    output_cells: list[str] = ['E2', 'E5', 'E7', 'H9', 'K10']
    calculation_cells: list[str] = ['H7', 'K7', 'K8']
    
    # Style cells
    style_cells(ws, heading_cells, "Headline 3")
    style_cells(ws, input_cells, 'Input')
    style_cells(ws, output_cells, 'Output')
    style_cells(ws, calculation_cells, 'Calculation')
    
    format_output_cells(ws)
    set_column_width(ws, ['A'], 7)
    set_column_width(ws, ['B'], 9)
    set_column_width(ws, ['D'], 11)
    set_column_width(ws, ['J'], 10)
    set_column_width(ws, ['G'], 5)
    
    set_all_borders(ws, "D1:E7")
    set_all_borders(ws, "G1:H9")
    set_all_borders(ws, "J1:K10")
    set_thick_outside_borders(ws, "G6:H6")
    set_thick_outside_borders(ws, "J5:K5")
    set_thick_outside_borders(ws, "J6:K6")
    
    add_functions(ws)
    readonly_cells: list = output_cells + calculation_cells + ['A1:B800']
    set_readonly_cells(ws, readonly_cells)
    
    # Save
    wb.save("z_table.xlsx")
    print("\nSuccessfully created z_table.xlsx\n")

if __name__ == "__main__":
    create_excel_file()
    